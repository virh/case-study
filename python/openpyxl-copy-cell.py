#!/usr/bin/env python
import sys, getopt, os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def main(argv):
	baseDir = ''
	sourceColumn = ''
	distColumn = ''
	try:
		opts, args = getopt.getopt(argv,"h:s:d:",[])
	except getopt.GetoptError:
		print('copy-cell.py -dir <dir> -s <sourcecolumn> -d <distcolumn>')
		sys.exit(2)
	for opt, arg in opts:
		if opt == '-h':
			print('copy-cell.py -dir <dir> -s <sourcecolumn> -d <distcolumn>')
			sys.exit()
		#elif opt == '-dir':
		#	baseDir = arg
		elif opt in ("-s", "--scolumn"):
			sourceColumn = arg
		elif opt in ("-d", "--dcolumn"):
			distColumn = arg

	print('Base dir is ', baseDir)
	print('Source column is ', sourceColumn)
	print('Dist column is ', distColumn)
	if (not sourceColumn) or (not distColumn):
		print("Usage: copy-cell.py -dir <dir> -s <sourcecolumn> -d <distcolumn>")
		return
	
	if not baseDir:
		basepath = '.'
	else:
		basepath = baseDir
	for entry in os.listdir(basepath):
		print(entry, ' ', 'xlsx ', os.path.splitext(entry)[1][1:])
		if os.path.isfile(os.path.join(basepath, entry)) and 'xlsx'==os.path.splitext(entry)[1][1:]:
			print(entry)
			#filename = os.path.join(basepath, entry)
			filename = "test1.xlsx"
			wb = load_workbook(filename)
			sheet = wb.active

			for row in range(1, sheet.max_row + 1):
				if not sheet[str(distColumn) + str(row)].value:
					sheet[str(distColumn) + str(row)] = sheet[str(sourceColumn) + str(row)].value

			wb.save(filename)
			print(filename)

if __name__ == "__main__":
   main(sys.argv[1:])